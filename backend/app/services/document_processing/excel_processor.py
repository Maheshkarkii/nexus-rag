"""Excel spreadsheet extractor supporting multi-sheet .xlsx workbooks using openpyxl."""

import logging
from pathlib import Path
from typing import List
import openpyxl
from app.db.models.document import Document
from app.services.document_processing.base import BaseDocumentProcessor, ExtractionResult

logger = logging.getLogger("ai_research_assistant.processors.excel")


class ExcelProcessor(BaseDocumentProcessor):
    """Extracts tabular worksheets from Microsoft Excel spreadsheets with macro safety."""

    async def extract(self, file_path: Path, document: Document) -> ExtractionResult:
        if not file_path.exists():
            raise FileNotFoundError(f"Physical Excel file not found at: {file_path}")

        try:
            # data_only=True extracts evaluated values and avoids evaluating VBA macros
            workbook = openpyxl.load_workbook(
                str(file_path), data_only=True, read_only=True, keep_vba=False
            )
        except Exception as e:
            raise ValueError(f"Failed to load Excel workbook: {e}") from e

        sheet_blocks: List[str] = []
        sheet_summaries = []

        try:
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                rows_data: List[List[str]] = []

                for row in sheet.iter_rows(values_only=True):
                    # Filter out purely empty rows
                    if row and any(cell is not None and str(cell).strip() for cell in row):
                        cleaned_cells = [
                            str(cell).strip() if cell is not None else "" for cell in row
                        ]
                        rows_data.append(cleaned_cells)

                if not rows_data:
                    continue

                header = rows_data[0]
                data_rows = rows_data[1:]

                sheet_text = [
                    f"Sheet: {sheet_name}",
                    f"Columns: {' | '.join(header)}",
                    "",
                ]

                for idx, row in enumerate(data_rows, start=1):
                    sheet_text.append(f"Row {idx}: {' | '.join(row)}")

                sheet_blocks.append("\n".join(sheet_text))
                sheet_summaries.append({
                    "sheet_name": sheet_name,
                    "row_count": len(data_rows),
                    "column_count": len(header),
                })
        finally:
            workbook.close()

        combined_text = "\n\n---\n\n".join(sheet_blocks).strip()

        if not combined_text:
            raise ValueError("Excel workbook contains no readable worksheet data or cell values.")

        char_count = len(combined_text)
        word_count = len(combined_text.split())

        return ExtractionResult(
            text=combined_text,
            character_count=char_count,
            word_count=word_count,
            metadata={
                "sheet_count": len(sheet_summaries),
                "sheets": sheet_summaries,
            },
        )
